import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
import shutil
import os

st.set_page_config(page_title="연차 관리 시스템", layout="wide")

st.title("📊 연차 사용 관리 프로그램")

FILE_PATH = "회사 연차사용.xlsx"
SHEET_NAME = "26년도 연차사용"
BACKUP_DIR = "backup"

# -----------------------------
# 공통 설정
# -----------------------------
USE_COLS = [f"사용일{i}" for i in range(1, 31)]


# -----------------------------
# 공통 함수
# -----------------------------
def get_target_year():
    return datetime.today().year


def to_number(value, default=0):
    num = pd.to_numeric(value, errors="coerce")
    return default if pd.isna(num) else float(num)


def clean_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def format_display_date(value):
    if pd.isna(value) or str(value).strip() == "" or str(value).strip().lower() == "none":
        return ""
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def format_leave_date(use_date, leave_type):
    date_str = pd.to_datetime(use_date).strftime("%Y-%m-%d")
    if leave_type == "반차":
        return f"{date_str} (반차)"
    return date_str


def parse_cancel_amount(value):
    text = str(value)
    return 0.5 if "반차" in text else 1.0


def calculate_service_years(hire_date, base_date):
    years = base_date.year - hire_date.year
    if (base_date.month, base_date.day) < (hire_date.month, hire_date.day):
        years -= 1
    return max(0, years)


def calculate_anniversary_period(hire_date, target_year):
    """
    해당 target_year의 기산시작일 / 기산종료일 계산
    예: 입사일 2023-05-10, target_year=2026
    -> 기산시작일 2026-05-10, 기산종료일 2027-05-09
    """
    try:
        start = date(target_year, hire_date.month, hire_date.day)
    except ValueError:
        if hire_date.month == 2 and hire_date.day == 29:
            start = date(target_year, 2, 28)
        else:
            raise

    try:
        end = date(target_year + 1, hire_date.month, hire_date.day) - timedelta(days=1)
    except ValueError:
        if hire_date.month == 2 and hire_date.day == 29:
            end = date(target_year + 1, 2, 28) - timedelta(days=1)
        else:
            raise

    return start, end


def calculate_auto_leave_days(hire_date, target_year=None):
    """
    입사일 기준 자동 연차 계산
    실무형 자동 계산:
    - 1년 미만: 월차 최대 11일
    - 1년 이상: 15일
    - 3년 이상: 2년마다 1일 가산
    - 최대 25일
    """
    if target_year is None:
        target_year = get_target_year()

    start_date, end_date = calculate_anniversary_period(hire_date, target_year)
    service_years = calculate_service_years(hire_date, start_date)

    # 1년 미만
    if service_years < 1:
        months_worked = (start_date.year - hire_date.year) * 12 + (start_date.month - hire_date.month)
        if start_date.day < hire_date.day:
            months_worked -= 1
        months_worked = max(0, min(11, months_worked))
        leave_days = float(months_worked)
    else:
        # 1년 이상 기본 15일
        # 3년 이상부터 2년마다 1일 가산
        extra_days = max(0, (service_years - 1) // 2)
        leave_days = float(min(25, 15 + extra_days))

    return start_date, end_date, service_years, leave_days


def find_first_empty_use_col(row):
    for col in USE_COLS:
        value = row.get(col, None)
        if pd.isna(value) or clean_text(value) == "" or clean_text(value).lower() == "none":
            return col
    return None


def style_remaining_leave(val):
    num = pd.to_numeric(val, errors="coerce")
    if pd.isna(num):
        return ""
    if num <= 0:
        return "background-color: #f8d7da; color: #842029; font-weight: bold;"
    elif num <= 5:
        return "background-color: #fff3cd; color: #664d03; font-weight: bold;"
    return ""


def format_leave_number(value):
    num = pd.to_numeric(value, errors="coerce")
    if pd.isna(num):
        return ""
    if float(num).is_integer():
        return str(int(num))
    return str(num)


def create_backup():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(BACKUP_DIR, f"회사 연차사용_backup_{timestamp}.xlsx")
    shutil.copy(FILE_PATH, backup_path)
    return backup_path


def parse_use_entry(value):
    """
    사용일 셀 값을 날짜와 사용량으로 해석
    예:
    2026-03-27 -> 날짜, 1.0
    2026-03-27 (반차) -> 날짜, 0.5
    2026.03.27 -> 날짜, 1.0
    """
    if pd.isna(value):
        return None, None

    text = str(value).strip()
    if text == "" or text.lower() == "none":
        return None, None

    amount = 0.5 if "반차" in text else 1.0
    clean = text.replace("(반차)", "").strip()

    parsed_date = pd.to_datetime(clean, errors="coerce")
    if pd.isna(parsed_date):
        return None, None

    return parsed_date, amount


def build_monthly_stats(df, target_year, target_month):
    rows = []
    total_count = 0
    total_amount = 0.0

    for _, row in df.iterrows():
        emp_name = str(row["이름"]).strip()
        emp_count = 0
        emp_amount = 0.0

        for col in USE_COLS:
            value = row.get(col, None)
            parsed_date, amount = parse_use_entry(value)
            if parsed_date is None:
                continue

            if parsed_date.year == target_year and parsed_date.month == target_month:
                emp_count += 1
                emp_amount += amount

        if emp_count > 0:
            rows.append({
                "이름": emp_name,
                "사용 건수": emp_count,
                "사용 일수": format_leave_number(emp_amount)
            })
            total_count += emp_count
            total_amount += emp_amount

    result_df = pd.DataFrame(rows)
    return result_df, total_count, total_amount


# -----------------------------
# 데이터 불러오기
# -----------------------------
@st.cache_data
def load_data():
    df = pd.read_excel(FILE_PATH, sheet_name=SHEET_NAME, header=1)
    df.columns = [str(c).strip() for c in df.columns]
    df = df[df["이름"].notna()].copy()

    for col in USE_COLS:
        if col not in df.columns:
            df[col] = None

    return df


def save_data_to_excel(df: pd.DataFrame):
    wb = load_workbook(FILE_PATH)
    ws = wb[SHEET_NAME]

    start_row = 3
    max_row = ws.max_row
    max_col = ws.max_column

    for r in range(start_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None

    for row_idx, row in enumerate(df.itertuples(index=False), start=start_row):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    wb.save(FILE_PATH)


df = load_data()

# -----------------------------
# 상단 관리 도구
# -----------------------------
st.subheader("🛠️ 관리 도구")

tool_col1, tool_col2 = st.columns(2)

with tool_col1:
    if st.button("💾 지금 백업하기", use_container_width=True):
        backup_file = create_backup()
        st.success(f"백업 완료: {backup_file}")

with tool_col2:
    if os.path.exists(BACKUP_DIR):
        backup_files = sorted(os.listdir(BACKUP_DIR), reverse=True)
        st.write(f"백업 파일 수: {len(backup_files)}")
    else:
        st.write("백업 파일 수: 0")

# -----------------------------
# 직원 검색 + 선택
# -----------------------------
st.subheader("👤 직원 선택")

names = sorted(df["이름"].dropna().astype(str).unique().tolist())

search_name = st.text_input("직원 검색", placeholder="이름을 입력하세요")
if search_name:
    filtered_names = [n for n in names if search_name.strip().lower() in n.lower()]
else:
    filtered_names = names

if not filtered_names:
    st.warning("검색 결과가 없습니다.")
    st.stop()

selected_name = st.selectbox("직원 선택", filtered_names)
employee = df[df["이름"] == selected_name].iloc[0]

# -----------------------------
# 현재 연차 정보
# -----------------------------
st.subheader("📌 현재 연차 현황")

col1, col2, col3 = st.columns(3)

total = to_number(employee["발생 연차"])
used = to_number(employee["사용 연차"])
remain = to_number(employee["잔여 연차"])

col1.metric("총 연차", format_leave_number(total))
col2.metric("사용 연차", format_leave_number(used))
col3.metric("잔여 연차", format_leave_number(remain))

if remain <= 0:
    st.error("잔여 연차가 없습니다.")
elif remain <= 5:
    st.warning("잔여 연차가 5일 이하입니다.")
else:
    st.success("잔여 연차가 충분합니다.")

# -----------------------------
# 연차 사용 입력
# -----------------------------
with st.expander("📝 연차 사용 입력", expanded=False):
    use_date = st.date_input("사용 날짜 선택", datetime.today())
    leave_type = st.radio("사용 종류 선택", ["연차", "반차"], horizontal=True)
    leave_amount = 1.0 if leave_type == "연차" else 0.5

    st.write(f"선택된 사용값: **{format_leave_number(leave_amount)}일**")

    btn_col1, btn_col2 = st.columns(2)

    with btn_col1:
        register_btn = st.button("등록하기", type="primary", use_container_width=True)

    with btn_col2:
        preview_btn = st.button("미리 확인", use_container_width=True)

    if preview_btn:
        expected_used = used + leave_amount
        expected_remain = total - expected_used
        st.info(
            f"{selected_name} / {use_date.strftime('%Y-%m-%d')} / {leave_type} 등록 시 "
            f"사용 연차 {format_leave_number(expected_used)}, 잔여 연차 {format_leave_number(expected_remain)}"
        )

    if register_btn:
        idx = df[df["이름"] == selected_name].index[0]

        current_total = to_number(df.loc[idx, "발생 연차"])
        current_used = to_number(df.loc[idx, "사용 연차"])
        current_remain = to_number(df.loc[idx, "잔여 연차"])

        if current_remain < leave_amount:
            st.error("잔여 연차가 부족합니다.")
        else:
            empty_col = find_first_empty_use_col(df.loc[idx])

            if empty_col is None:
                st.error("사용일 칸이 모두 찼습니다. 사용일1~사용일30을 확인해주세요.")
            else:
                df.loc[idx, empty_col] = format_leave_date(use_date, leave_type)
                df.loc[idx, "사용 연차"] = current_used + leave_amount
                df.loc[idx, "잔여 연차"] = current_total - (current_used + leave_amount)

                save_data_to_excel(df)
                st.cache_data.clear()
                st.success(f"{leave_type} 등록 완료!")
                st.rerun()

# -----------------------------
# 선택 직원 사용일 보기
# -----------------------------
with st.expander("🗂️ 선택 직원 사용일 내역", expanded=False):
    use_list = []
    for col in USE_COLS:
        value = employee.get(col, None)
        if pd.notna(value) and clean_text(value) != "" and clean_text(value).lower() != "none":
            use_list.append({"구분": col, "사용내역": format_display_date(value)})

    use_df = pd.DataFrame(use_list)

    if not use_df.empty:
        st.dataframe(use_df, use_container_width=True)
    else:
        st.info("등록된 사용일이 없습니다.")

# -----------------------------
# 연차 취소 기능
# -----------------------------
with st.expander("↩️ 연차 취소", expanded=False):
    if not use_df.empty:
        cancel_options = [
            f"{row['구분']} | {row['사용내역']}"
            for _, row in use_df.iterrows()
        ]
        selected_cancel = st.selectbox("취소할 사용일 선택", cancel_options)

        if st.button("선택 사용일 취소", use_container_width=True):
            idx = df[df["이름"] == selected_name].index[0]

            selected_col = selected_cancel.split("|")[0].strip()
            selected_value = df.loc[idx, selected_col]

            cancel_amount = parse_cancel_amount(selected_value)

            current_total = to_number(df.loc[idx, "발생 연차"])
            current_used = to_number(df.loc[idx, "사용 연차"])

            new_used = max(0, current_used - cancel_amount)
            new_remain = current_total - new_used

            df.loc[idx, selected_col] = None
            df.loc[idx, "사용 연차"] = new_used
            df.loc[idx, "잔여 연차"] = new_remain

            save_data_to_excel(df)
            st.cache_data.clear()
            st.success("연차 취소 완료!")
            st.rerun()
    else:
        st.info("취소할 사용일이 없습니다.")

# -----------------------------
# 직원 관리
# -----------------------------
with st.expander("📁 직원 관리", expanded=False):
    # -------------------------
    # 직원 추가
    # -------------------------
    st.markdown("## ➕ 직원 추가")

    with st.form("add_employee_form"):
        new_name = st.text_input("직원 이름")
        new_hire_date = st.date_input("입사일", value=date.today())

        preview_start, preview_end, preview_service_years, preview_leave_days = calculate_auto_leave_days(
            new_hire_date,
            get_target_year()
        )

        st.info(
            f"자동 계산 결과\n\n"
            f"- 기산시작일: {preview_start}\n"
            f"- 기산종료일: {preview_end}\n"
            f"- 근속년수: {preview_service_years}\n"
            f"- 발생 연차: {format_leave_number(preview_leave_days)}일"
        )

        submit_add_employee = st.form_submit_button("직원 추가하기")

        if submit_add_employee:
            new_name = new_name.strip()

            if new_name == "":
                st.error("직원 이름을 입력해주세요.")
            elif new_name in df["이름"].astype(str).tolist():
                st.error("이미 등록된 직원입니다.")
            else:
                hire_date = pd.to_datetime(new_hire_date).date()
                start_date, end_date, service_years, auto_leave_days = calculate_auto_leave_days(
                    hire_date,
                    get_target_year()
                )

                new_row = {
                    "이름": new_name,
                    "입사일": pd.to_datetime(hire_date),
                    "기산시작일": pd.to_datetime(start_date),
                    "기산종료일": pd.to_datetime(end_date),
                    "근속년수\n(기산일 기준)": service_years,
                    "발생 연차": float(auto_leave_days),
                    "사용 연차": 0.0,
                    "잔여 연차": float(auto_leave_days),
                }

                for col in USE_COLS:
                    new_row[col] = None

                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

                save_data_to_excel(df)
                st.cache_data.clear()
                st.success(f"{new_name} 직원 추가 완료!")
                st.rerun()

    st.markdown("---")

    # -------------------------
    # 직원 수정
    # -------------------------
    st.markdown("## ✏️ 직원 수정")

    edit_name = st.selectbox("수정할 직원 선택", names, key="edit_employee_select")
    edit_employee = df[df["이름"] == edit_name].iloc[0]

    default_hire_date = pd.to_datetime(edit_employee["입사일"], errors="coerce")
    if pd.isna(default_hire_date):
        default_hire_date = pd.Timestamp(date.today())

    with st.form("edit_employee_form"):
        edited_name = st.text_input("직원 이름 수정", value=str(edit_employee["이름"]))
        edited_hire_date = st.date_input("입사일 수정", value=default_hire_date.date())
        edited_used_leave = st.number_input(
            "사용 연차 수정",
            min_value=0.0,
            step=0.5,
            value=float(to_number(edit_employee["사용 연차"]))
        )

        preview_start, preview_end, preview_service_years, preview_leave_days = calculate_auto_leave_days(
            pd.to_datetime(edited_hire_date).date(),
            get_target_year()
        )

        st.info(
            f"자동 계산 결과\n\n"
            f"- 기산시작일: {preview_start}\n"
            f"- 기산종료일: {preview_end}\n"
            f"- 근속년수: {preview_service_years}\n"
            f"- 발생 연차: {format_leave_number(preview_leave_days)}일"
        )

        submit_edit_employee = st.form_submit_button("직원 정보 수정하기")

        if submit_edit_employee:
            edited_name = edited_name.strip()

            if edited_name == "":
                st.error("직원 이름을 입력해주세요.")
            else:
                duplicate_names = [n for n in df["이름"].astype(str).tolist() if n != edit_name]
                if edited_name in duplicate_names:
                    st.error("같은 이름의 직원이 이미 있습니다.")
                else:
                    idx = df[df["이름"] == edit_name].index[0]

                    hire_date = pd.to_datetime(edited_hire_date).date()
                    start_date, end_date, service_years, auto_leave_days = calculate_auto_leave_days(
                        hire_date,
                        get_target_year()
                    )

                    new_total = float(auto_leave_days)
                    new_used = float(edited_used_leave)
                    new_remain = new_total - new_used

                    if new_remain < 0:
                        st.error("사용 연차가 발생 연차보다 클 수 없습니다.")
                    else:
                        df.loc[idx, "이름"] = edited_name
                        df.loc[idx, "입사일"] = pd.to_datetime(hire_date)
                        df.loc[idx, "기산시작일"] = pd.to_datetime(start_date)
                        df.loc[idx, "기산종료일"] = pd.to_datetime(end_date)
                        df.loc[idx, "근속년수\n(기산일 기준)"] = service_years
                        df.loc[idx, "발생 연차"] = new_total
                        df.loc[idx, "사용 연차"] = new_used
                        df.loc[idx, "잔여 연차"] = new_remain

                        save_data_to_excel(df)
                        st.cache_data.clear()
                        st.success(f"{edited_name} 직원 정보 수정 완료!")
                        st.rerun()

    st.markdown("---")

    # -------------------------
    # 직원 삭제
    # -------------------------
    st.markdown("## 🗑️ 직원 삭제")

    delete_name = st.selectbox("삭제할 직원 선택", names, key="delete_employee_select")
    confirm_delete = st.checkbox("정말 삭제합니다. 되돌리기 어렵습니다.")

    if st.button("선택 직원 삭제", use_container_width=True):
        if not confirm_delete:
            st.warning("삭제 확인 체크를 먼저 해주세요.")
        else:
            before_count = len(df)
            df = df[df["이름"].astype(str) != str(delete_name)].copy()
            after_count = len(df)

            if before_count == after_count:
                st.error("삭제할 직원을 찾지 못했습니다.")
            else:
                save_data_to_excel(df)
                st.cache_data.clear()
                st.success(f"{delete_name} 직원 삭제 완료!")
                st.rerun()

# -----------------------------
# 월별 통계
# -----------------------------
with st.expander("📅 월별 연차 통계", expanded=False):
    stat_col1, stat_col2 = st.columns(2)

    with stat_col1:
        stat_year = st.number_input("조회 연도", min_value=2020, max_value=2100, value=get_target_year(), step=1)

    with stat_col2:
        stat_month = st.selectbox("조회 월", list(range(1, 13)), index=max(0, datetime.today().month - 1))

    monthly_df, monthly_count, monthly_amount = build_monthly_stats(df, int(stat_year), int(stat_month))

    metric_col1, metric_col2 = st.columns(2)
    metric_col1.metric("해당 월 사용 건수", monthly_count)
    metric_col2.metric("해당 월 총 사용일수", format_leave_number(monthly_amount))

    if not monthly_df.empty:
        st.dataframe(monthly_df, use_container_width=True)
    else:
        st.info("해당 월 사용 내역이 없습니다.")

# -----------------------------
# 전체 현황 보기
# -----------------------------
with st.expander("📋 전체 연차 현황", expanded=False):
    show_cols = [
        "이름", "입사일", "기산시작일", "기산종료일", "근속년수\n(기산일 기준)",
        "발생 연차", "사용 연차", "잔여 연차",
        "사용일1", "사용일2", "사용일3", "사용일4", "사용일5", "사용일6", "사용일7"
    ]

    display_df = df.copy()

    for col in ["입사일", "기산시작일", "기산종료일"]:
        if col in display_df.columns:
            display_df[col] = pd.to_datetime(display_df[col], errors="coerce").dt.strftime("%Y-%m-%d")

    for col in USE_COLS:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(format_display_date)

    display_df["발생 연차"] = display_df["발생 연차"].apply(format_leave_number)
    display_df["사용 연차"] = display_df["사용 연차"].apply(format_leave_number)
    display_df["잔여 연차"] = display_df["잔여 연차"].apply(format_leave_number)

    existing_cols = [col for col in show_cols if col in display_df.columns]
    styled_df = display_df[existing_cols].style.map(style_remaining_leave, subset=["잔여 연차"])
    st.dataframe(styled_df, use_container_width=True)

# -----------------------------
# 엑셀 다운로드
# -----------------------------
with st.expander("⬇️ 엑셀 다운로드", expanded=False):
    with open(FILE_PATH, "rb") as f:
        st.download_button(
            label="엑셀 다운로드",
            data=f,
            file_name="회사 연차사용.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )