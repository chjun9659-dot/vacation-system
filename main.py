import streamlit as st
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
import shutil
import os
import io
import pickle
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

# 👉 여기에 넣으세요
DRIVE_SCOPES = ['https://www.googleapis.com/auth/drive.file']

st.set_page_config(page_title="윤우 통합 운영 시스템", layout="wide")

# =========================================================
# 0. 로그인 설정
# =========================================================
USERS = {
    "admin": {"password": "1234", "role": "관리자"},
    "staff": {"password": "1234", "role": "직원"},
    "시공": {"password": "1234", "role": "시공"},
    "행정": {"password": "1234", "role": "관리자"},
}

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if "username" not in st.session_state:
    st.session_state.username = ""

if "role" not in st.session_state:
    st.session_state.role = ""

if "menu" not in st.session_state:
    st.session_state.menu = "홈"

if "inspection_form_version" not in st.session_state:
    st.session_state.inspection_form_version = 0

if "inspection_edit_mode" not in st.session_state:
    st.session_state.inspection_edit_mode = False

if "inspection_edit_target" not in st.session_state:
    st.session_state.inspection_edit_target = None        


def logout():
    st.session_state.logged_in = False
    st.session_state.username = ""
    st.session_state.role = ""
    st.session_state.menu = "홈"
    st.rerun()


def login_screen():
    st.title("🔐 윤우 통합 운영 시스템 로그인")
    st.write("연차관리 + 시공일정 + 실사관리를 하나로 사용하는 통합 시스템입니다.")

    with st.form("login_form_unique"):
        username = st.text_input("아이디")
        password = st.text_input("비밀번호", type="password")
        submitted = st.form_submit_button("로그인")

        if submitted:
            username = username.strip()
            if username in USERS and USERS[username]["password"] == password:
                st.session_state.logged_in = True
                st.session_state.username = username
                st.session_state.role = USERS[username]["role"]
                st.success("로그인 성공!")
                st.rerun()
            else:
                st.error("아이디 또는 비밀번호가 올바르지 않습니다.")


# =========================================================
# 1. 공통 UI
# =========================================================
def draw_sidebar():
    menu_list = ["홈", "연차 관리", "시공 일정", "실사 관리"]

    with st.sidebar:
        st.markdown(f"### 👤 {st.session_state.username}")
        st.caption(f"권한: {st.session_state.role}")

        if st.session_state.menu not in menu_list:
            st.session_state.menu = "홈"

        selected_menu = st.radio(
            "메뉴 선택",
            menu_list,
            index=menu_list.index(st.session_state.menu)
        )

        if selected_menu != st.session_state.menu:
            st.session_state.menu = selected_menu
            st.rerun()

        st.divider()

        if st.button("로그아웃", use_container_width=True, key="logout_btn_unique"):
            logout()


def home_page():
    if st.session_state.menu != "홈":
        return
    st.markdown("""
    <style>
    .home-title {
        font-size: 28px;
        font-weight: 800;
        color: #0f172a;
        margin-bottom: 6px;
    }
    .home-desc {
        font-size: 14px;
        color: #475569;
        margin-bottom: 18px;
    }
    .home-guide-title {
        font-size: 18px;
        font-weight: 700;
        color: #0f172a;
        margin-top: 10px;
        margin-bottom: 8px;
    }
    .home-guide-box {
        border: 1px solid #e5e7eb;
        border-radius: 12px;
        background: #ffffff;
        padding: 16px 18px;
        line-height: 1.8;
        color: #0f172a;
        font-size: 15px;
    }
    div.stButton > button {
        border-radius: 12px !important;
        min-height: 72px !important;
        font-size: 20px !important;
        font-weight: 700 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown('<div class="home-title">🏢 윤우 통합 운영 시스템</div>', unsafe_allow_html=True)
    st.markdown('<div class="home-desc">회사용 내부 운영 시스템입니다.</div>', unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)

    with c1:
        if st.button("📊 연차 관리", use_container_width=True, key="home_btn_vacation"):
            st.session_state.menu = "연차 관리"
            st.rerun()
            return

    with c2:
        if st.button("📅 시공 일정", use_container_width=True, key="home_btn_schedule"):
            st.session_state.menu = "시공 일정"
            st.rerun()
            return

    with c3:
        if st.button("🔎 실사 관리", use_container_width=True, key="home_btn_inspection"):
            st.session_state.menu = "실사 관리"
            st.rerun()
            return

    st.divider()

    st.markdown('<div class="home-guide-title">빠른 안내</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="home-guide-box">
    • 좌측 메뉴에서 원하는 시스템을 선택하세요.<br>
    • 메인 화면 버튼을 눌러도 각 시스템으로 바로 이동할 수 있습니다.<br>
    • 현재는 하나의 앱에서 연차 / 시공일정 / 실사관리를 함께 사용할 수 있습니다.
    </div>
    """, unsafe_allow_html=True)


# =========================================================
# 2. 연차 관리 시스템
# =========================================================
VACATION_FILE_PATH = "회사 연차사용.xlsx"
VACATION_SHEET_NAME = "26년도 연차사용"
VACATION_BACKUP_DIR = "backup"
USE_COLS = [f"사용일{i}" for i in range(1, 31)]


def get_target_year():
    return datetime.today().year


def to_number(value, default=0):
    num = pd.to_numeric(value, errors="coerce")
    return default if pd.isna(num) else float(num)


def clean_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def format_display_date(value):
    if pd.isna(value) or str(value).strip() == "" or str(value).strip().lower() == "none":
        return ""
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def format_leave_date(use_date, leave_type):
    date_str = pd.to_datetime(use_date).strftime("%Y-%m-%d")
    if leave_type == "반차":
        return f"{date_str} (반차)"
    return date_str


def parse_cancel_amount(value):
    text = str(value)
    return 0.5 if "반차" in text else 1.0


def calculate_service_years(hire_date, base_date):
    years = base_date.year - hire_date.year
    if (base_date.month, base_date.day) < (hire_date.month, hire_date.day):
        years -= 1
    return max(0, years)


def calculate_anniversary_period(hire_date, target_year):
    try:
        start = date(target_year, hire_date.month, hire_date.day)
    except ValueError:
        if hire_date.month == 2 and hire_date.day == 29:
            start = date(target_year, 2, 28)
        else:
            raise

    try:
        end = date(target_year + 1, hire_date.month, hire_date.day) - timedelta(days=1)
    except ValueError:
        if hire_date.month == 2 and hire_date.day == 29:
            end = date(target_year + 1, 2, 28) - timedelta(days=1)
        else:
            raise

    return start, end


def calculate_auto_leave_days(hire_date, target_year=None):
    if target_year is None:
        target_year = get_target_year()

    start_date, end_date = calculate_anniversary_period(hire_date, target_year)
    service_years = calculate_service_years(hire_date, start_date)

    if service_years < 1:
        months_worked = (start_date.year - hire_date.year) * 12 + (start_date.month - hire_date.month)
        if start_date.day < hire_date.day:
            months_worked -= 1
        months_worked = max(0, min(11, months_worked))
        leave_days = float(months_worked)
    else:
        extra_days = max(0, (service_years - 1) // 2)
        leave_days = float(min(25, 15 + extra_days))

    return start_date, end_date, service_years, leave_days


def find_first_empty_use_col(row, df_columns):
    for col in USE_COLS:
        matching_indexes = [i for i, c in enumerate(df_columns) if str(c).strip() == col]

        for col_idx in matching_indexes:
            value = row.iloc[col_idx]

            if pd.isna(value) or clean_text(value) == "" or clean_text(value).lower() == "none":
                return col_idx

    return None


def style_remaining_leave(val):
    num = pd.to_numeric(val, errors="coerce")
    if pd.isna(num):
        return ""
    if num <= 0:
        return "background-color: #f8d7da; color: #842029; font-weight: bold;"
    elif num <= 5:
        return "background-color: #fff3cd; color: #664d03; font-weight: bold;"
    return ""


def format_leave_number(value):
    num = pd.to_numeric(value, errors="coerce")
    if pd.isna(num):
        return ""
    if float(num).is_integer():
        return str(int(num))
    return str(num)


def create_backup():
    os.makedirs(VACATION_BACKUP_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(VACATION_BACKUP_DIR, f"회사 연차사용_backup_{timestamp}.xlsx")
    shutil.copy(VACATION_FILE_PATH, backup_path)
    return backup_path


def parse_use_entry(value):
    if pd.isna(value):
        return None, None

    text = str(value).strip()
    if text == "" or text.lower() == "none":
        return None, None

    amount = 0.5 if "반차" in text else 1.0
    clean = text.replace("(반차)", "").strip()

    parsed_date = pd.to_datetime(clean, errors="coerce")
    if pd.isna(parsed_date):
        return None, None

    return parsed_date, amount

def recalculate_vacation_summary(df: pd.DataFrame):
    for idx in df.index:
        total_leave = to_number(df.loc[idx, "발생 연차"])
        used_leave = 0.0

        for col in USE_COLS:
            if col not in df.columns:
                continue

            value = df.loc[idx, col]

            if pd.isna(value):
                continue

            text = str(value).strip()
            if text == "" or text.lower() == "none":
                continue

            if "반차" in text:
                used_leave += 0.5
            else:
                used_leave += 1.0

        remain_leave = total_leave - used_leave

        df.loc[idx, "사용 연차"] = used_leave
        df.loc[idx, "잔여 연차"] = remain_leave

    return df

def build_monthly_stats(df, target_year, target_month):
    rows = []
    total_count = 0
    total_amount = 0.0

    for _, row in df.iterrows():
        emp_name = str(row["이름"]).strip()
        emp_count = 0
        emp_amount = 0.0

        for col in USE_COLS:
            value = row.get(col, None)
            parsed_date, amount = parse_use_entry(value)
            if parsed_date is None:
                continue

            if parsed_date.year == target_year and parsed_date.month == target_month:
                emp_count += 1
                emp_amount += amount

        if emp_count > 0:
            rows.append({
                "이름": emp_name,
                "사용 건수": emp_count,
                "사용 일수": format_leave_number(emp_amount)
            })
            total_count += emp_count
            total_amount += emp_amount

    result_df = pd.DataFrame(rows)
    return result_df, total_count, total_amount


@st.cache_data(ttl=60)
def load_vacation_data():
    df = pd.read_excel(VACATION_FILE_PATH, sheet_name=VACATION_SHEET_NAME, header=1)
    df.columns = [str(c).strip() for c in df.columns]

    # 중복 컬럼 제거
    df = df.loc[:, ~df.columns.duplicated()].copy()

    df = df[df["이름"].notna()].copy()

    for col in USE_COLS:
        if col not in df.columns:
            df[col] = None

    # ✅ 사용일 컬럼은 문자열 저장 가능하도록 object 고정
    for col in USE_COLS:
        df[col] = df[col].astype("object")

    # ✅ 반차(0.5) 저장 가능하도록 연차 수치 컬럼은 float 고정
    number_cols = ["발생 연차", "사용 연차", "잔여 연차"]
    for col in number_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(float)

    df = recalculate_vacation_summary(df)
    return df


def save_vacation_data_to_excel(df: pd.DataFrame):
    df = df.copy()

    # ✅ 숫자 컬럼 강제 float 처리 (반차 0.5 대응)
    for col in ["발생 연차", "사용 연차", "잔여 연차"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(float)

    wb = load_workbook(VACATION_FILE_PATH)
    ws = wb[VACATION_SHEET_NAME]

    header_row = 2     # 실제 헤더 행
    start_row = 3      # 실제 데이터 시작 행

    # 엑셀 헤더 읽기
    excel_headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        excel_headers.append(str(v).strip() if v is not None else "")

    # 헤더명 -> 엑셀 컬럼번호
    col_map = {name: idx for idx, name in enumerate(excel_headers, start=1) if name}

    # df에 있는 컬럼만 엑셀에서 지우기
    last_row = max(ws.max_row, start_row + len(df) + 50)
    for r in range(start_row, last_row + 1):
        for col_name in df.columns:
            if col_name in col_map:
                ws.cell(row=r, column=col_map[col_name]).value = None

    # 헤더명 기준으로 정확히 저장
    for row_idx, (_, row) in enumerate(df.iterrows(), start=start_row):
        for col_name in df.columns:
            if col_name in col_map:
                value = row[col_name]

                if pd.isna(value):
                    value = None

                ws.cell(row=row_idx, column=col_map[col_name]).value = value

    wb.save(VACATION_FILE_PATH)


def vacation_page():
    render_inspection_common_style()

    st.markdown('<div class="erp-page-title">윤우테크 연차 관리 프로그램</div>', unsafe_allow_html=True)
    st.markdown('<div class="erp-page-desc">연차 관리 및 현황 확인 프로그램입니다.</div>', unsafe_allow_html=True)

    try:
        df = load_vacation_data()
    except Exception as e:
        st.error(f"연차 파일을 불러오지 못했습니다: {e}")
        return

    st.subheader("🛠️ 관리 도구")

    tool_col1, tool_col2, tool_col3 = st.columns(3)

    with tool_col1:
        if st.button("💾 지금 백업하기", use_container_width=True, key="vac_backup_btn_unique"):
            backup_file = create_backup()
            st.success(f"백업 완료: {backup_file}")

    with tool_col2:
        if st.button("🧮 연차 수치 재정리", use_container_width=True, key="vac_recalc_btn_unique"):
            df = recalculate_vacation_summary(df)
            save_vacation_data_to_excel(df)
            st.cache_data.clear()
            st.success("사용일 기준으로 사용 연차 / 잔여 연차를 재정리했습니다.")
            st.rerun()

    with tool_col3:
        if os.path.exists(VACATION_BACKUP_DIR):
            backup_files = sorted(os.listdir(VACATION_BACKUP_DIR), reverse=True)
            st.write(f"백업 파일 수: {len(backup_files)}")
        else:
            st.write("백업 파일 수: 0")

    st.subheader("👤 직원 선택")

    names = sorted(df["이름"].dropna().astype(str).unique().tolist())

    search_name = st.text_input("직원 검색", placeholder="이름을 입력하세요", key="vac_search_name_unique")
    if search_name:
        filtered_names = [n for n in names if search_name.strip().lower() in n.lower()]
    else:
        filtered_names = names

    if not filtered_names:
        st.warning("검색 결과가 없습니다.")
        return

    selected_name = st.selectbox("직원 선택", filtered_names, key="vac_selected_name_unique")
    employee = df[df["이름"] == selected_name].iloc[0]

    st.subheader("📌 현재 연차 현황")

    col1, col2, col3 = st.columns(3)

    total = to_number(employee["발생 연차"])
    used = to_number(employee["사용 연차"])
    remain = to_number(employee["잔여 연차"])

    col1.metric("총 연차", format_leave_number(total))
    col2.metric("사용 연차", format_leave_number(used))
    col3.metric("잔여 연차", format_leave_number(remain))

    if remain <= 0:
        st.error("잔여 연차가 없습니다.")
    elif remain <= 5:
        st.warning("잔여 연차가 5일 이하입니다.")
    else:
        st.success("잔여 연차가 충분합니다.")

    with st.expander("📝 연차 사용 입력", expanded=False):
        use_date = st.date_input("사용 날짜 선택", datetime.today(), key="vac_use_date_unique")
        leave_type = st.radio("사용 종류 선택", ["연차", "반차"], horizontal=True, key="vac_leave_type_unique")
        leave_amount = 1.0 if leave_type == "연차" else 0.5

        st.write(f"선택된 사용값: **{format_leave_number(leave_amount)}일**")

        btn_col1, btn_col2 = st.columns(2)

        with btn_col1:
            register_btn = st.button("등록하기", type="primary", use_container_width=True, key="vac_register_btn_unique")

        with btn_col2:
            preview_btn = st.button("미리 확인", use_container_width=True, key="vac_preview_btn_unique")

        if preview_btn:
            expected_used = used + leave_amount
            expected_remain = total - expected_used
            st.info(
                f"{selected_name} / {use_date.strftime('%Y-%m-%d')} / {leave_type} 등록 시 "
                f"사용 연차 {format_leave_number(expected_used)}, 잔여 연차 {format_leave_number(expected_remain)}"
            )

        if register_btn:
            idx = df[df["이름"] == selected_name].index[0]

            current_total = float(to_number(df.loc[idx, "발생 연차"]))
            current_used = float(to_number(df.loc[idx, "사용 연차"]))
            current_remain = float(to_number(df.loc[idx, "잔여 연차"]))

            if current_remain < leave_amount:
                st.error("잔여 연차가 부족합니다.")
            else:
                row_pos = df.index.get_loc(idx)
                empty_col_idx = find_first_empty_use_col(df.iloc[row_pos], df.columns)

                if empty_col_idx is None:
                    st.error("사용일 칸이 모두 찼습니다. 사용일1~사용일30을 확인해주세요.")
                else:
                    df.iat[row_pos, empty_col_idx] = format_leave_date(use_date, leave_type)

                    df.loc[idx, "사용 연차"] = float(current_used + leave_amount)
                    df.loc[idx, "잔여 연차"] = float(current_total - (current_used + leave_amount))

                    save_vacation_data_to_excel(df)
                    st.cache_data.clear()
                    st.success(f"{leave_type} 등록 완료!")
                    st.rerun()

    with st.expander("🗂️ 선택 직원 사용일 내역", expanded=False):
        use_list = []
        for col in USE_COLS:
            value = employee.get(col, None)
            if pd.notna(value) and clean_text(value) != "" and clean_text(value).lower() != "none":
                use_list.append({"구분": col, "사용내역": format_display_date(value)})

        use_df = pd.DataFrame(use_list)

        if not use_df.empty:
            st.dataframe(use_df, use_container_width=True)
        else:
            st.info("등록된 사용일이 없습니다.")

    with st.expander("↩️ 연차 취소", expanded=False):
        use_list = []
        for col in USE_COLS:
            value = employee.get(col, None)
            if pd.notna(value) and clean_text(value) != "" and clean_text(value).lower() != "none":
                use_list.append({"구분": col, "사용내역": format_display_date(value)})

        use_df = pd.DataFrame(use_list)

        if not use_df.empty:
            cancel_options = [f"{row['구분']} | {row['사용내역']}" for _, row in use_df.iterrows()]
            selected_cancel = st.selectbox("취소할 사용일 선택", cancel_options, key="vac_cancel_select_unique")

            if st.button("선택 사용일 취소", use_container_width=True, key="vac_cancel_btn_unique"):
                idx = df[df["이름"] == selected_name].index[0]

                selected_col = selected_cancel.split("|")[0].strip()
                selected_value = df.loc[idx, selected_col]

                cancel_amount = parse_cancel_amount(selected_value)

                current_total = to_number(df.loc[idx, "발생 연차"])
                current_used = to_number(df.loc[idx, "사용 연차"])

                new_used = max(0, current_used - cancel_amount)
                new_remain = current_total - new_used

                df.loc[idx, selected_col] = None
                df.loc[idx, "사용 연차"] = new_used
                df.loc[idx, "잔여 연차"] = new_remain

                save_vacation_data_to_excel(df)
                st.cache_data.clear()
                st.success("연차 취소 완료!")
                st.rerun()
        else:
            st.info("취소할 사용일이 없습니다.")

    with st.expander("📁 직원 관리", expanded=False):
        st.markdown("## ➕ 직원 추가")

        with st.form("add_employee_form_unique"):
            new_name = st.text_input("직원 이름", key="new_employee_name_unique")
            new_hire_date = st.date_input("입사일", value=date.today(), key="new_employee_hire_date_unique")

            preview_start, preview_end, preview_service_years, preview_leave_days = calculate_auto_leave_days(
                new_hire_date,
                get_target_year()
            )

            st.info(
                f"자동 계산 결과\n\n"
                f"- 기산시작일: {preview_start}\n"
                f"- 기산종료일: {preview_end}\n"
                f"- 근속년수: {preview_service_years}\n"
                f"- 발생 연차: {format_leave_number(preview_leave_days)}일"
            )

            submit_add_employee = st.form_submit_button("직원 추가하기")

            if submit_add_employee:
                new_name = new_name.strip()

                if new_name == "":
                    st.error("직원 이름을 입력해주세요.")
                elif new_name in df["이름"].astype(str).tolist():
                    st.error("이미 등록된 직원입니다.")
                else:
                    hire_date = pd.to_datetime(new_hire_date).date()
                    start_date, end_date, service_years, auto_leave_days = calculate_auto_leave_days(
                        hire_date,
                        get_target_year()
                    )

                    new_row = {
                        "이름": new_name,
                        "입사일": pd.to_datetime(hire_date),
                        "기산시작일": pd.to_datetime(start_date),
                        "기산종료일": pd.to_datetime(end_date),
                        "근속년수\n(기산일 기준)": service_years,
                        "발생 연차": float(auto_leave_days),
                        "사용 연차": 0.0,
                        "잔여 연차": float(auto_leave_days),
                    }

                    for col in USE_COLS:
                        new_row[col] = None

                    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

                    save_vacation_data_to_excel(df)
                    st.cache_data.clear()
                    st.success(f"{new_name} 직원 추가 완료!")
                    st.rerun()

        st.markdown("---")
        st.markdown("## ✏️ 직원 수정")

        edit_name = st.selectbox("수정할 직원 선택", names, key="edit_employee_select_unique")
        edit_employee = df[df["이름"] == edit_name].iloc[0]

        default_hire_date = pd.to_datetime(edit_employee["입사일"], errors="coerce")
        if pd.isna(default_hire_date):
            default_hire_date = pd.Timestamp(date.today())

        with st.form("edit_employee_form_unique"):
            edited_name = st.text_input("직원 이름 수정", value=str(edit_employee["이름"]), key="edited_name_unique")
            edited_hire_date = st.date_input("입사일 수정", value=default_hire_date.date(), key="edited_hire_date_unique")
            edited_used_leave = st.number_input(
                "사용 연차 수정",
                min_value=0.0,
                step=0.5,
                value=float(to_number(edit_employee["사용 연차"])),
                key="edited_used_leave_unique"
            )

            preview_start, preview_end, preview_service_years, preview_leave_days = calculate_auto_leave_days(
                pd.to_datetime(edited_hire_date).date(),
                get_target_year()
            )

            st.info(
                f"자동 계산 결과\n\n"
                f"- 기산시작일: {preview_start}\n"
                f"- 기산종료일: {preview_end}\n"
                f"- 근속년수: {preview_service_years}\n"
                f"- 발생 연차: {format_leave_number(preview_leave_days)}일"
            )

            submit_edit_employee = st.form_submit_button("직원 정보 수정하기")

            if submit_edit_employee:
                edited_name = edited_name.strip()

                if edited_name == "":
                    st.error("직원 이름을 입력해주세요.")
                else:
                    duplicate_names = [n for n in df["이름"].astype(str).tolist() if n != edit_name]
                    if edited_name in duplicate_names:
                        st.error("같은 이름의 직원이 이미 있습니다.")
                    else:
                        idx = df[df["이름"] == edit_name].index[0]

                        hire_date = pd.to_datetime(edited_hire_date).date()
                        start_date, end_date, service_years, auto_leave_days = calculate_auto_leave_days(
                            hire_date,
                            get_target_year()
                        )

                        new_total = float(auto_leave_days)
                        new_used = float(edited_used_leave)
                        new_remain = new_total - new_used

                        if new_remain < 0:
                            st.error("사용 연차가 발생 연차보다 클 수 없습니다.")
                        else:
                            df.loc[idx, "이름"] = edited_name
                            df.loc[idx, "입사일"] = pd.to_datetime(hire_date)
                            df.loc[idx, "기산시작일"] = pd.to_datetime(start_date)
                            df.loc[idx, "기산종료일"] = pd.to_datetime(end_date)
                            df.loc[idx, "근속년수\n(기산일 기준)"] = service_years
                            df.loc[idx, "발생 연차"] = new_total
                            df.loc[idx, "사용 연차"] = new_used
                            df.loc[idx, "잔여 연차"] = new_remain

                            save_vacation_data_to_excel(df)
                            st.cache_data.clear()
                            st.success(f"{edited_name} 직원 정보 수정 완료!")
                            st.rerun()

        st.markdown("---")
        st.markdown("## 🗑️ 직원 삭제")

        delete_name = st.selectbox("삭제할 직원 선택", names, key="delete_employee_select_unique")
        confirm_delete = st.checkbox("정말 삭제합니다. 되돌리기 어렵습니다.", key="vac_confirm_delete_unique")

        if st.button("선택 직원 삭제", use_container_width=True, key="vac_delete_btn_unique"):
            if not confirm_delete:
                st.warning("삭제 확인 체크를 먼저 해주세요.")
            else:
                before_count = len(df)
                df = df[df["이름"].astype(str) != str(delete_name)].copy()
                after_count = len(df)

                if before_count == after_count:
                    st.error("삭제할 직원을 찾지 못했습니다.")
                else:
                    save_vacation_data_to_excel(df)
                    st.cache_data.clear()
                    st.success(f"{delete_name} 직원 삭제 완료!")
                    st.rerun()

    with st.expander("📅 월별 연차 통계", expanded=False):
        stat_col1, stat_col2 = st.columns(2)

        with stat_col1:
            stat_year = st.number_input(
                "조회 연도",
                min_value=2020,
                max_value=2100,
                value=get_target_year(),
                step=1,
                key="vac_stat_year_unique"
            )

        with stat_col2:
            stat_month = st.selectbox(
                "조회 월",
                list(range(1, 13)),
                index=max(0, datetime.today().month - 1),
                key="vac_stat_month_unique"
            )

        monthly_df, monthly_count, monthly_amount = build_monthly_stats(df, int(stat_year), int(stat_month))

        metric_col1, metric_col2 = st.columns(2)
        metric_col1.metric("해당 월 사용 건수", monthly_count)
        metric_col2.metric("해당 월 총 사용일수", format_leave_number(monthly_amount))

        if not monthly_df.empty:
            st.dataframe(monthly_df, use_container_width=True)
        else:
            st.info("해당 월 사용 내역이 없습니다.")

    with st.expander("📋 전체 연차 현황", expanded=False):
        show_cols = [
            "이름", "입사일", "기산시작일", "기산종료일", "근속년수\n(기산일 기준)",
            "발생 연차", "사용 연차", "잔여 연차",
            "사용일1", "사용일2", "사용일3", "사용일4", "사용일5", "사용일6", "사용일7"
        ]

        display_df = df.copy()

        for col in ["입사일", "기산시작일", "기산종료일"]:
            if col in display_df.columns:
                display_df[col] = pd.to_datetime(display_df[col], errors="coerce").dt.strftime("%Y-%m-%d")

        for col in USE_COLS:
            if col in display_df.columns:
                display_df[col] = display_df[col].apply(format_display_date)

        display_df["발생 연차"] = display_df["발생 연차"].apply(format_leave_number)
        display_df["사용 연차"] = display_df["사용 연차"].apply(format_leave_number)
        display_df["잔여 연차"] = display_df["잔여 연차"].apply(format_leave_number)

        existing_cols = [col for col in show_cols if col in display_df.columns]
        styled_df = display_df[existing_cols].style.map(style_remaining_leave, subset=["잔여 연차"])
        st.dataframe(styled_df, use_container_width=True)

    with st.expander("⬇️ 엑셀 다운로드", expanded=False):
        with open(VACATION_FILE_PATH, "rb") as f:
            st.download_button(
                label="엑셀 다운로드",
                data=f,
                file_name="회사 연차사용.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="vac_download_btn_unique"
            )


# =========================================================
# 3. 구글시트 / 구글드라이브 공통 설정
# =========================================================
# =========================================================
# 3. 구글시트 / 구글드라이브 공통 설정
# =========================================================
SCOPE = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/drive"
]

DRIVE_SCOPES = ['https://www.googleapis.com/auth/drive.file']


@st.cache_resource
def get_gspread_client():
    try:
        if os.path.exists("key.json"):
            creds = ServiceAccountCredentials.from_json_keyfile_name("key.json", SCOPE)
            return gspread.authorize(creds)

        secrets_dict = st.secrets.to_dict()
        if "gcp_service_account" in secrets_dict:
            creds_dict = secrets_dict["gcp_service_account"]
            creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, SCOPE)
            return gspread.authorize(creds)

        raise FileNotFoundError("key.json 또는 st.secrets의 gcp_service_account를 찾지 못했습니다.")

    except Exception as e:
        raise Exception(f"구글시트 인증 실패: {e}")


@st.cache_resource
def get_drive_service_oauth():
    creds = None

    base_dir = os.path.dirname(os.path.abspath(__file__))
    token_path = os.path.join(base_dir, "token.pickle")
    client_secret_path = os.path.join(base_dir, "client_secret.json")

    if not os.path.exists(client_secret_path):
        raise Exception(f"client_secret.json 파일을 찾지 못했습니다: {client_secret_path}")

    if os.path.exists(token_path):
        with open(token_path, "rb") as token:
            creds = pickle.load(token)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                client_secret_path,
                DRIVE_SCOPES
            )
            creds = flow.run_local_server(port=0)

        with open(token_path, "wb") as token:
            pickle.dump(creds, token)

    return build("drive", "v3", credentials=creds)


def upload_file_to_drive(uploaded_file, folder_id=None):
    try:
        drive_service = get_drive_service_oauth()

        file_stream = io.BytesIO(uploaded_file.getvalue())
        file_metadata = {"name": uploaded_file.name}

        if folder_id:
            file_metadata["parents"] = [folder_id]

        media = MediaIoBaseUpload(
            file_stream,
            mimetype=uploaded_file.type,
            resumable=False
        )

        uploaded = drive_service.files().create(
            body=file_metadata,
            media_body=media,
            fields="id, name, webViewLink"
        ).execute()

        file_id = uploaded.get("id")

        # 첨부파일 열람 권한 부여
        drive_service.permissions().create(
            fileId=file_id,
            body={
                "type": "anyone",
                "role": "reader"
            }
        ).execute()

        return uploaded.get("name", uploaded_file.name), uploaded.get("webViewLink", "")

    except Exception as e:
        raise Exception(f"파일 업로드 실패: {e}")


# =========================================================
# 4. 시공 일정 시스템
# =========================================================
SCHEDULE_SHEET_NAME = "시공일정"
EXPECTED_COLUMNS = ["날짜", "설치현장", "시공담당", "수량", "비고", "상태", "완료일"]


def get_schedule_sheet():
    client = get_gspread_client()
    return client.open(SCHEDULE_SHEET_NAME).sheet1


def ensure_schedule_sheet_header(sheet):
    values = sheet.get_all_values()
    if not values:
        sheet.update("A1:G1", [EXPECTED_COLUMNS])
        return

    header = values[0]
    if header != EXPECTED_COLUMNS:
        existing = pd.DataFrame(values[1:], columns=header if header else None)
        for col in EXPECTED_COLUMNS:
            if col not in existing.columns:
                existing[col] = ""
        existing = existing[EXPECTED_COLUMNS]
        save_schedule_data(existing, sheet)

@st.cache_data(ttl=60)
def load_schedule_data():
    sheet = get_schedule_sheet()
    ensure_schedule_sheet_header(sheet)

    data = sheet.get_all_records()
    df = pd.DataFrame(data)

    if df.empty:
        return pd.DataFrame(columns=EXPECTED_COLUMNS)

    for col in EXPECTED_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    df = df[EXPECTED_COLUMNS]

    df["수량"] = pd.to_numeric(df["수량"], errors="coerce").fillna(0).astype(int)
    df["날짜"] = df["날짜"].astype(str)
    df["완료일"] = df["완료일"].astype(str)
    df["상태"] = df["상태"].astype(str).replace("", "진행중")

    return df


def save_schedule_data(df, sheet=None):
    if sheet is None:
        sheet = get_schedule_sheet()

    save_df = df.copy()
    for col in EXPECTED_COLUMNS:
        if col not in save_df.columns:
            save_df[col] = ""

    save_df = save_df[EXPECTED_COLUMNS].fillna("")
    save_df["수량"] = pd.to_numeric(save_df["수량"], errors="coerce").fillna(0).astype(int)

    rows = [save_df.columns.tolist()] + save_df.astype(str).values.tolist()
    sheet.clear()
    sheet.update(rows)


def schedule_page():
    render_inspection_common_style()

    st.markdown('<div class="erp-page-title">시공 일정 관리 프로그램</div>', unsafe_allow_html=True)
    st.markdown('<div class="erp-page-desc">시공 일정 등록, 수정, 진행 현황 관리</div>', unsafe_allow_html=True)

    try:
        df = load_schedule_data()
    except Exception as e:
        st.error(f"시공일정 데이터를 불러오지 못했습니다: {e}")
        return

    df = df.reset_index(drop=True)
    df["row_id"] = df.index

    today_str = str(date.today())

    total_count = len(df)
    today_count = len(df[df["날짜"] == today_str])
    progress_count = len(df[df["상태"] == "진행중"])
    done_count = len(df[df["상태"] == "완료"])
    total_qty = int(df["수량"].sum()) if not df.empty else 0

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("전체 일정", total_count)
    c2.metric("오늘 일정", today_count)
    c3.metric("진행중", progress_count)
    c4.metric("완료", done_count)
    c5.metric("총 수량", total_qty)

    st.divider()

    st.subheader("1. 시공 일정 등록")

    with st.form("add_schedule_form_unique"):
        a1, a2, a3 = st.columns(3)
        work_date = a1.date_input("시공 날짜", value=date.today(), key="sch_work_date_unique")
        site_name = a2.text_input("설치현장", key="sch_site_name_unique")
        manager_name = a3.text_input("시공담당", key="sch_manager_name_unique")

        a4, a5 = st.columns(2)
        quantity = a4.number_input("수량", min_value=0, step=1, value=0, key="sch_quantity_unique")
        note = a5.text_input("비고", key="sch_note_unique")

        submitted = st.form_submit_button("등록하기")

        if submitted:
            if not site_name.strip():
                st.warning("설치현장을 입력해주세요.")
            elif not manager_name.strip():
                st.warning("시공담당을 입력해주세요.")
            else:
                new_row = pd.DataFrame([{
                    "날짜": str(work_date),
                    "설치현장": site_name.strip(),
                    "시공담당": manager_name.strip(),
                    "수량": int(quantity),
                    "비고": note.strip(),
                    "상태": "진행중",
                    "완료일": ""
                }])

                save_df = df[EXPECTED_COLUMNS].copy() if not df.empty else pd.DataFrame(columns=EXPECTED_COLUMNS)
                save_df = pd.concat([save_df, new_row], ignore_index=True)
                save_schedule_data(save_df)
                st.success("등록 완료!")
                st.rerun()

    st.divider()
    with st.expander("📅 2. 오늘 일정", expanded=False):
        today_df = df[df["날짜"] == today_str].copy()

        if today_df.empty:
            st.info("오늘 일정이 없습니다.")
        else:
            show_today = today_df[["날짜", "설치현장", "시공담당", "수량", "비고", "상태", "완료일"]]
            st.dataframe(show_today, use_container_width=True, hide_index=True)

    st.divider()
    with st.expander("📋 3. 시공 일정 보기", expanded=False):
        managers = ["전체"] + sorted([m for m in df["시공담당"].dropna().unique().tolist() if str(m).strip() != ""])

        f1, f2, f3, f4 = st.columns(4)
        status_filter = f1.selectbox("상태 선택", ["전체", "진행중", "완료"], key="sch_status_filter_unique")
        manager_filter = f2.selectbox("담당자 선택", managers, key="sch_manager_filter_unique")
        date_filter = f3.selectbox("날짜 기준", ["전체", "오늘", "미래", "지난 일정"], key="sch_date_filter_unique")
        keyword = f4.text_input("검색", placeholder="설치현장 / 비고 검색", key="sch_keyword_unique")

        filtered_df = df.copy()

        if status_filter != "전체":
            filtered_df = filtered_df[filtered_df["상태"] == status_filter]

        if manager_filter != "전체":
            filtered_df = filtered_df[filtered_df["시공담당"] == manager_filter]

        if date_filter == "오늘":
            filtered_df = filtered_df[filtered_df["날짜"] == today_str]
        elif date_filter == "미래":
            filtered_df = filtered_df[filtered_df["날짜"] > today_str]
        elif date_filter == "지난 일정":
            filtered_df = filtered_df[filtered_df["날짜"] < today_str]

        if keyword.strip():
            kw = keyword.strip()
            filtered_df = filtered_df[
                filtered_df["설치현장"].astype(str).str.contains(kw, case=False, na=False) |
                filtered_df["비고"].astype(str).str.contains(kw, case=False, na=False)
            ]

        show_df = filtered_df[["날짜", "설치현장", "시공담당", "수량", "비고", "상태", "완료일"]].copy()

        if show_df.empty:
            st.info("조건에 맞는 일정이 없습니다.")
        else:
            st.dataframe(show_df, use_container_width=True, hide_index=True)

    st.divider()

    with st.expander("✏️ 4. 일정 수정", expanded=False):
        if df.empty:
            st.info("수정할 일정이 없습니다.")
        else:
            edit_options = [
                f"{row['row_id']} | {row['날짜']} | {row['설치현장']} | {row['시공담당']}"
                for _, row in df.iterrows()
            ]

            selected_edit = st.selectbox("수정할 일정 선택", edit_options, key="sch_edit_select_unique")
            edit_idx = int(selected_edit.split("|")[0].strip())
            edit_row = df.loc[df["row_id"] == edit_idx].iloc[0]

            default_edit_date = (
                pd.to_datetime(edit_row["날짜"]).date()
                if str(edit_row["날짜"]).strip()
                else date.today()
            )

            with st.form(f"edit_schedule_form_{edit_idx}"):
                e1, e2, e3 = st.columns(3)
                edit_date = e1.date_input("시공 날짜 수정", value=default_edit_date)
                edit_site = e2.text_input("설치현장 수정", value=str(edit_row["설치현장"]))
                edit_manager = e3.text_input("시공담당 수정", value=str(edit_row["시공담당"]))

                e4, e5, e6 = st.columns(3)
                edit_qty = e4.number_input("수량 수정", min_value=0, step=1, value=int(edit_row["수량"]))
                edit_note = e5.text_input("비고 수정", value=str(edit_row["비고"]))
                edit_status = e6.selectbox(
                    "상태 수정",
                    ["진행중", "완료"],
                    index=0 if str(edit_row["상태"]) == "진행중" else 1
                )

                edit_submit = st.form_submit_button("수정 저장")

                if edit_submit:
                    save_df = df[EXPECTED_COLUMNS].copy()
                    save_df.loc[edit_idx, "날짜"] = str(edit_date)
                    save_df.loc[edit_idx, "설치현장"] = edit_site.strip()
                    save_df.loc[edit_idx, "시공담당"] = edit_manager.strip()
                    save_df.loc[edit_idx, "수량"] = int(edit_qty)
                    save_df.loc[edit_idx, "비고"] = edit_note.strip()
                    save_df.loc[edit_idx, "상태"] = edit_status

                    if edit_status == "완료" and not str(save_df.loc[edit_idx, "완료일"]).strip():
                        save_df.loc[edit_idx, "완료일"] = today_str
                    elif edit_status == "진행중":
                        save_df.loc[edit_idx, "완료일"] = ""

                    save_schedule_data(save_df)
                    st.success("수정 완료!")
                    st.rerun()

    st.divider()

    with st.expander("✅ 5. 완료 처리", expanded=False):
        progress_df = df[df["상태"] == "진행중"].copy()

        if progress_df.empty:
            st.info("완료 처리할 일정이 없습니다.")
        else:
            complete_options = [
                f"{row['row_id']} | {row['날짜']} | {row['설치현장']} | {row['시공담당']} | 수량 {row['수량']}"
                for _, row in progress_df.iterrows()
            ]

            selected_complete = st.selectbox("완료 처리 일정 선택", complete_options, key="sch_complete_select_unique")

            if st.button("완료로 변경", key="sch_complete_btn_unique"):
                complete_idx = int(selected_complete.split("|")[0].strip())
                save_df = df[EXPECTED_COLUMNS].copy()
                save_df.loc[complete_idx, "상태"] = "완료"
                save_df.loc[complete_idx, "완료일"] = today_str
                save_schedule_data(save_df)
                st.success("완료 처리되었습니다.")
                st.rerun()

    st.divider()

    with st.expander("↩️ 6. 완료 취소", expanded=False):
        done_df = df[df["상태"] == "완료"].copy()

        if done_df.empty:
            st.info("완료 취소할 일정이 없습니다.")
        else:
            cancel_options = [
                f"{row['row_id']} | {row['날짜']} | {row['설치현장']} | {row['시공담당']} | 수량 {row['수량']}"
                for _, row in done_df.iterrows()
            ]

            selected_cancel = st.selectbox("완료 취소 일정 선택", cancel_options, key="sch_cancel_select_unique")

            if st.button("진행중으로 변경", key="sch_cancel_btn_unique"):
                cancel_idx = int(selected_cancel.split("|")[0].strip())
                save_df = df[EXPECTED_COLUMNS].copy()
                save_df.loc[cancel_idx, "상태"] = "진행중"
                save_df.loc[cancel_idx, "완료일"] = ""
                save_schedule_data(save_df)
                st.success("완료 취소되었습니다.")
                st.rerun()

    st.divider()

    with st.expander("🗑️ 7. 일정 삭제", expanded=False):
        if df.empty:
            st.info("삭제할 일정이 없습니다.")
        else:
            delete_options = [
                f"{row['row_id']} | {row['날짜']} | {row['설치현장']} | {row['시공담당']} | 수량 {row['수량']}"
                for _, row in df.iterrows()
            ]

            selected_delete = st.selectbox("삭제할 일정 선택", delete_options, key="sch_delete_select_unique")

            if st.button("선택 일정 삭제", key="sch_delete_btn_unique"):
                delete_idx = int(selected_delete.split("|")[0].strip())
                save_df = df[EXPECTED_COLUMNS].copy()
                save_df = save_df.drop(index=delete_idx).reset_index(drop=True)
                save_schedule_data(save_df)
                st.success("삭제 완료!")
                st.rerun()


# =========================================================
# 5. 실사 관리 시스템
# =========================================================
INSPECTION_SHEET_NAME = "실사관리"
INSPECTION_COLUMNS = [
    "요청일",
    "운영사",
    "현장명",
    "현장주소",
    "현장연락처",
    "주차면수",
    "상품구분",
    "신규설치수량",
    "기설치수량",
    "영업담당자",
    "영업담당연락처",
    "요청내용",
    "비고",
    "첨부파일명",
    "첨부파일링크",
    "실사담당자",
    "실사예정일",
    "실사완료일",
    "진행상태",
    "실사결과",
    "특이사항",
    "후속조치",
    "계약여부",
    "계약일",
    "계약수량",
    "계약금액",
    "미계약사유"
]

INSPECTION_STATUS_OPTIONS = [
    "요청접수",
    "담당자배정",
    "일정확정",
    "실사진행",
    "실사완료",
    "계약완료",
    "미계약종결"
]

PRODUCT_OPTIONS = ["아이센서", "전기차충전기", "이전설치"]
CONTRACT_OPTIONS = ["대기", "계약", "미계약"]


def get_inspection_sheet():
    client = get_gspread_client()
    spreadsheet = client.open(INSPECTION_SHEET_NAME)
    worksheet = spreadsheet.worksheet("실사복구")
    return worksheet


def safe_int(value, default=0):
    num = pd.to_numeric(value, errors="coerce")
    if pd.isna(num):
        return default
    return int(num)


def show_inspection_flash():
    msg = st.session_state.pop("inspection_flash", "")
    msg_type = st.session_state.pop("inspection_flash_type", "success")

    if msg:
        if msg_type == "success":
            st.success(msg)
        elif msg_type == "warning":
            st.warning(msg)
        elif msg_type == "error":
            st.error(msg)
        else:
            st.info(msg)


def set_inspection_flash(msg, msg_type="success"):
    st.session_state["inspection_flash"] = msg
    st.session_state["inspection_flash_type"] = msg_type


def normalize_inspection_df(df):
    for col in INSPECTION_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    df = df[INSPECTION_COLUMNS].copy()

    int_cols = ["주차면수", "신규설치수량", "기설치수량", "계약수량"]
    for col in int_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    df["계약금액"] = pd.to_numeric(df["계약금액"], errors="coerce").fillna(0)

    date_cols = ["요청일", "실사예정일", "실사완료일", "계약일"]
    for col in date_cols:
        df[col] = df[col].astype(str)

    df["진행상태"] = df["진행상태"].astype(str).replace("", "요청접수")
    df["계약여부"] = df["계약여부"].astype(str).replace("", "대기")

    return df

@st.cache_data(ttl=60)
def load_inspection_data():
    try:
        sheet = get_inspection_sheet()
        values = sheet.get_all_values()

        if not values:
            return pd.DataFrame(columns=INSPECTION_COLUMNS)

        header = values[0]
        data_rows = values[1:]

        # 👉 기존 데이터 그대로 로드
        temp_df = pd.DataFrame(data_rows, columns=header)

        # 👉 새로운 DF 생성
        df = pd.DataFrame(columns=INSPECTION_COLUMNS)

        # 👉 컬럼 매핑 (안전 버전)
        column_map = {
            "요청일": "요청일",
            "운영사": "운영사",
            "현장명": "현장명",
            "단지명": "현장명",
            "현장주소": "현장주소",
            "주소": "현장주소",
            "현장연락처": "현장연락처",
            "전화번호": "현장연락처",
            "주차면수": "주차면수",
            "상품구분": "상품구분",
            "신규설치수량": "신규설치수량",
            "수량": "신규설치수량",
            "기설치수량": "기설치수량",
            "영업담당자": "영업담당자",
            "영업담당연락처": "영업담당연락처",
            "요청내용": "요청내용",
            "비고": "비고",
            "첨부파일명": "첨부파일명",
            "첨부파일링크": "첨부파일링크",
            "실사담당자": "실사담당자",
            "실사예정일": "실사예정일",
            "실사완료일": "실사완료일",
            "진행상태": "진행상태",
            "실사결과": "실사결과",
            "특이사항": "특이사항",
            "후속조치": "후속조치",
            "계약여부": "계약여부",
            "계약일": "계약일",
            "계약수량": "계약수량",
            "계약금액": "계약금액",
            "미계약사유": "미계약사유",
        }

        # 👉 안전하게 매핑
        for old_col, new_col in column_map.items():
            if old_col in temp_df.columns:
                df[new_col] = temp_df[old_col]

        # 👉 없는 컬럼 채우기
        for col in INSPECTION_COLUMNS:
            if col not in df.columns:
                df[col] = ""

        return df

    except Exception as e:
        st.error(f"실사 데이터를 불러오지 못했습니다: {e}")
        return pd.DataFrame(columns=INSPECTION_COLUMNS)


def save_inspection_data(df, sheet=None):
    if sheet is None:
        sheet = get_inspection_sheet()

    save_df = normalize_inspection_df(df)
    rows = [save_df.columns.tolist()] + save_df.astype(str).values.tolist()
    sheet.clear()
    sheet.update(rows)

def render_inspection_common_style():
    st.markdown("""
    <style>
    .erp-page-title {
    font-size: 28px !important;
    font-weight: 700 !important;
    color: #0f172a !important;
    margin-bottom: 4px !important;
}

.erp-page-desc {
    font-size: 14px !important;
    color: #64748b !important;
    margin-bottom: 20px !important;
}

    .erp-summary-card {
        border: 1px solid #e5e7eb;
        border-radius: 12px;
        background: #ffffff;
        padding: 12px 14px;
        min-height: 88px;
    }

    .erp-summary-label {
        font-size: 12px;
        color: #64748b;
        margin-bottom: 10px;
    }

    .erp-summary-value {
        font-size: 22px;
        font-weight: 700;
        color: #0f172a;
        line-height: 1.2;
    }

    .erp-section-title {
        font-size: 18px;
        font-weight: 700;
        color: #0f172a;
        margin-bottom: 10px;
    }

    .erp-label {
        font-size: 13px;
        font-weight: 600;
        color: #334155;
        margin-bottom: 6px;
    }

    .erp-box {
        border: 1px solid #dbe3ee;
        border-radius: 10px;
        background: #eef5ff;
        padding: 11px 13px;
        min-height: 44px;
        font-size: 14px;
        color: #0f172a;
        display: flex;
        align-items: center;
    }

    .erp-soft-box {
        border: 1px solid #e5e7eb;
        border-radius: 10px;
        background: #f8fafc;
        padding: 14px;
        font-size: 14px;
        color: #0f172a;
        line-height: 1.6;
    }

    .erp-divider {
        margin: 14px 0 18px 0;
    }

    div[data-testid="stExpander"] {
        border: 1px solid #e5e7eb !important;
        border-radius: 12px !important;
        background: #ffffff !important;
    }

    div[data-testid="stExpander"] details summary p {
        font-size: 16px !important;
        font-weight: 700 !important;
        color: #0f172a !important;
    }

    div[data-testid="stDataFrame"] {
        border-radius: 12px;
        overflow: hidden;
    }

    div[data-testid="stForm"] {
        border: 1px solid #e5e7eb;
        border-radius: 12px;
        background: #ffffff;
        padding: 14px 14px 6px 14px;
    }

    button[kind="primary"] {
        border-radius: 10px !important;
    }

    button[kind="secondary"] {
        border-radius: 10px !important;
    }
    </style>
    """, unsafe_allow_html=True)

def inspection_page():
    render_inspection_common_style()

    st.markdown('<div class="erp-page-title">🔎 실사 관리 프로그램</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="erp-page-desc">실사 요청 등록 → 담당자 배정 → 일정 입력 → 결과 작성 → 계약 여부 관리</div>',
        unsafe_allow_html=True
    )

    show_inspection_flash()

    df = load_inspection_data()
    df = df.reset_index(drop=True)
    df["row_id"] = df.index

    total_count = len(df)
    pending_count = len(df[df["진행상태"] == "요청접수"])
    assigned_count = len(df[df["진행상태"].isin(["담당자배정", "일정확정", "실사진행"])])
    done_count = len(df[df["진행상태"] == "실사완료"])
    contract_done_count = len(df[df["계약여부"] == "계약"])

    c1, c2, c3, c4, c5 = st.columns(5)

    with c1:
        st.markdown(
            f"""
            <div class="erp-summary-card">
                <div class="erp-summary-label">전체 요청</div>
                <div class="erp-summary-value">{total_count}</div>
            </div>
            """,
            unsafe_allow_html=True
        )

    with c2:
        st.markdown(
            f"""
            <div class="erp-summary-card">
                <div class="erp-summary-label">요청접수</div>
                <div class="erp-summary-value">{pending_count}</div>
            </div>
            """,
            unsafe_allow_html=True
        )

    with c3:
        st.markdown(
            f"""
            <div class="erp-summary-card">
                <div class="erp-summary-label">진행중</div>
                <div class="erp-summary-value">{assigned_count}</div>
            </div>
            """,
            unsafe_allow_html=True
        )

    with c4:
        st.markdown(
            f"""
            <div class="erp-summary-card">
                <div class="erp-summary-label">실사완료</div>
                <div class="erp-summary-value">{done_count}</div>
            </div>
            """,
            unsafe_allow_html=True
        )

    with c5:
        st.markdown(
            f"""
            <div class="erp-summary-card">
                <div class="erp-summary-label">계약완료</div>
                <div class="erp-summary-value">{contract_done_count}</div>
            </div>
            """,
            unsafe_allow_html=True
        )

    st.divider()

    with st.expander("📝 1. 실사 요청 등록", expanded=False):
        st.markdown('<div class="erp-section-title">실사 요청 등록</div>', unsafe_allow_html=True)
        form_ver = st.session_state.inspection_form_version

        with st.form(f"inspection_request_form_new_{form_ver}"):
            c1, c2, c3 = st.columns(3)
            req_date = c1.date_input("요청일", value=date.today(), key=f"req_date_{form_ver}")
            operator_name = c2.text_input("운영사", key=f"operator_name_{form_ver}")
            site_name = c3.text_input("현장명", key=f"site_name_{form_ver}")

            c4, c5, c6 = st.columns(3)
            site_address = c4.text_input("현장주소", key=f"site_address_{form_ver}")
            site_phone = c5.text_input("현장연락처", key=f"site_phone_{form_ver}")
            product_type = c6.selectbox("상품구분", PRODUCT_OPTIONS, key=f"product_type_{form_ver}")

            c7, c8, c9 = st.columns(3)
            parking_count = c7.number_input("주차면수", min_value=0, step=1, value=0, key=f"parking_count_{form_ver}")
            new_qty = c8.number_input("신규설치수량", min_value=0, step=1, value=0, key=f"new_qty_{form_ver}")
            installed_qty = c9.number_input("기설치수량", min_value=0, step=1, value=0, key=f"installed_qty_{form_ver}")

            c10, c11 = st.columns(2)
            sales_manager = c10.text_input("영업담당자", key=f"sales_manager_{form_ver}")
            sales_phone = c11.text_input("영업담당자 연락처", key=f"sales_phone_{form_ver}")

            request_content = st.text_area("요청내용", key=f"request_content_{form_ver}")
            note = st.text_input("비고", key=f"note_{form_ver}")

            st.markdown("#### 첨부파일")
            uploaded_file = st.file_uploader(
                "실사 관련 파일 업로드",
                type=["pdf", "png", "jpg", "jpeg", "xlsx", "xls", "doc", "docx"],
                key=f"insp_uploaded_file_new_{form_ver}"
            )

            submit_request = st.form_submit_button("실사 요청 등록")

            if submit_request:
                if not site_name.strip():
                    st.warning("현장명을 입력해주세요.")
                elif not sales_manager.strip():
                    st.warning("영업담당자를 입력해주세요.")
                else:
                    attachment_name = ""
                    attachment_link = ""

                    if uploaded_file is not None:
                        try:
                            attachment_name, attachment_link = upload_file_to_drive(
                                uploaded_file,
                                folder_id="1_TVqakggj2P-0ZnVLgEyCqjiqnxAf-nr"
                            )
                        except Exception as e:
                            st.error(str(e))
                            attachment_name = ""
                            attachment_link = ""

                    new_row = pd.DataFrame([{
                        "요청일": str(req_date),
                        "운영사": operator_name.strip(),
                        "현장명": site_name.strip(),
                        "현장주소": site_address.strip(),
                        "현장연락처": site_phone.strip(),
                        "주차면수": int(parking_count),
                        "상품구분": product_type,
                        "신규설치수량": int(new_qty),
                        "기설치수량": int(installed_qty),
                        "영업담당자": sales_manager.strip(),
                        "영업담당연락처": sales_phone.strip(),
                        "요청내용": request_content.strip(),
                        "비고": note.strip(),
                        "첨부파일명": attachment_name,
                        "첨부파일링크": attachment_link,
                        "실사담당자": "",
                        "실사예정일": "",
                        "실사완료일": "",
                        "진행상태": "요청접수",
                        "실사결과": "",
                        "특이사항": "",
                        "후속조치": "",
                        "계약여부": "대기",
                        "계약일": "",
                        "계약수량": 0,
                        "계약금액": 0,
                        "미계약사유": ""
                    }])

                    save_df = df[INSPECTION_COLUMNS].copy() if not df.empty else pd.DataFrame(columns=INSPECTION_COLUMNS)
                    save_df = pd.concat([save_df, new_row], ignore_index=True)
                    save_inspection_data(save_df)

                    set_inspection_flash("실사 요청이 등록되었습니다.", "success")
                    st.session_state.inspection_form_version += 1
                    st.rerun()

    st.divider()

    with st.expander("📋 2. 전체 실사 현황", expanded=False):
        st.markdown('<div class="erp-section-title">전체 실사 현황</div>', unsafe_allow_html=True)
        st.markdown('<div class="erp-soft-box">전체 실사 데이터를 확인합니다.</div>', unsafe_allow_html=True)

        status_list = ["전체"] + INSPECTION_STATUS_OPTIONS
        product_list = ["전체"] + PRODUCT_OPTIONS
        contract_list = ["전체"] + CONTRACT_OPTIONS

        f1, f2, f3, f4 = st.columns(4)
        status_filter = f1.selectbox("진행상태", status_list, key="insp_filter_status_new")
        product_filter = f2.selectbox("상품구분", product_list, key="insp_filter_product_new")
        contract_filter = f3.selectbox("계약여부", contract_list, key="insp_filter_contract_new")
        keyword = f4.text_input("검색", placeholder="현장명 / 주소 / 담당자 / 운영사", key="insp_filter_keyword_new")

        filtered_df = df.copy()

        if status_filter != "전체":
            filtered_df = filtered_df[filtered_df["진행상태"] == status_filter]

        if product_filter != "전체":
            filtered_df = filtered_df[filtered_df["상품구분"] == product_filter]

        if contract_filter != "전체":
            filtered_df = filtered_df[filtered_df["계약여부"] == contract_filter]

        if keyword.strip():
            kw = keyword.strip()
            filtered_df = filtered_df[
                filtered_df["현장명"].astype(str).str.contains(kw, case=False, na=False) |
                filtered_df["현장주소"].astype(str).str.contains(kw, case=False, na=False) |
                filtered_df["영업담당자"].astype(str).str.contains(kw, case=False, na=False) |
                filtered_df["실사담당자"].astype(str).str.contains(kw, case=False, na=False) |
                filtered_df["운영사"].astype(str).str.contains(kw, case=False, na=False)
            ].copy()

        show_df = filtered_df[[
            "요청일",
            "상품구분",
            "현장명",
            "현장주소",
            "현장연락처",
            "운영사",
            "신규설치수량",
            "기설치수량",
            "주차면수",

            "영업담당자",
            "영업담당연락처",

            "실사담당자",
            "실사예정일",
            "진행상태",
            "계약여부",

            "첨부파일링크"
        ]].copy()

        def status_style(val):
            if str(val) == "요청접수":
                return "background-color: #fef3c7; color: #92400e; font-weight: 600;"
            elif str(val) in ["담당자배정", "일정확정", "실사진행"]:
                return "background-color: #dbeafe; color: #1d4ed8; font-weight: 600;"
            elif str(val) == "실사완료":
                return "background-color: #dcfce7; color: #166534; font-weight: 600;"
            elif str(val) == "계약완료":
                return "background-color: #dcfce7; color: #166534; font-weight: 700;"
            elif str(val) == "미계약종결":
                return "background-color: #fee2e2; color: #991b1b; font-weight: 600;"
            return ""

        def contract_style(val):
            if str(val) == "계약":
                return "background-color: #dcfce7; color: #166534; font-weight: 700;"
            elif str(val) == "미계약":
                return "background-color: #fee2e2; color: #991b1b; font-weight: 700;"
            elif str(val) == "대기":
                return "background-color: #f3f4f6; color: #374151; font-weight: 600;"
            return ""

        if show_df.empty:
            st.info("조건에 맞는 실사 내역이 없습니다.")
        else:
            show_df["첨부파일열기"] = show_df["첨부파일링크"]

            styled_df = show_df.style.map(status_style, subset=["진행상태"]).map(contract_style, subset=["계약여부"])

            st.dataframe(
                styled_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "첨부파일링크": None,
                    "첨부파일열기": st.column_config.LinkColumn(
                        "첨부파일 열기",
                        display_text="열기"
                    )
                }
            )

            st.caption("상태와 계약여부는 색상으로 구분되며, 첨부파일은 '열기'로 바로 확인할 수 있습니다.")

    st.divider()

    with st.expander("🧑‍🔧 3. 담당자 배정 / 일정 입력", expanded=False):
        if df.empty:
            st.info("배정할 실사 요청이 없습니다.")
        else:
            assign_options = [
                f"{row['row_id']} | {row['요청일']} | {row['현장명']} | {row['상품구분']} | {row['영업담당자']}"
                for _, row in df.iterrows()
            ]

            selected_assign = st.selectbox(
                "배정할 실사 선택",
                assign_options,
                key="insp_assign_select_new"
            )
            assign_idx = int(selected_assign.split("|")[0].strip())
            assign_row = df.loc[df["row_id"] == assign_idx].iloc[0]

            assign_date_raw = str(assign_row["실사예정일"]).strip()
            parsed_assign_date = pd.to_datetime(assign_date_raw, errors="coerce")
            default_assign_date = parsed_assign_date.date() if pd.notna(parsed_assign_date) else date.today()

            with st.form(f"inspection_assign_form_{assign_idx}"):
                a1, a2, a3 = st.columns(3)

                inspector = a1.text_input(
                    "실사담당자",
                    value=str(assign_row["실사담당자"])
                )

                inspect_date = a2.date_input(
                    "실사예정일",
                    value=default_assign_date
                )

                current_status = str(assign_row["진행상태"]).strip()
                default_status_index = (
                    INSPECTION_STATUS_OPTIONS.index(current_status)
                    if current_status in INSPECTION_STATUS_OPTIONS
                    else 0
                )

                inspect_status = a3.selectbox(
                    "진행상태",
                    INSPECTION_STATUS_OPTIONS,
                    index=default_status_index
                )

                b1, b2 = st.columns(2)
                assign_submit = b1.form_submit_button("배정 / 일정 저장")
                delete_submit = b2.form_submit_button("담당자 배정 삭제")

                if assign_submit:
                    save_df = df[INSPECTION_COLUMNS].copy()
                    save_df.loc[assign_idx, "실사담당자"] = inspector.strip()
                    save_df.loc[assign_idx, "실사예정일"] = str(inspect_date)
                    save_df.loc[assign_idx, "진행상태"] = inspect_status
                    save_inspection_data(save_df)
                    st.cache_data.clear()

                    set_inspection_flash("담당자 배정 및 일정 저장 완료!", "success")
                    st.rerun()

                if delete_submit:
                    save_df = df[INSPECTION_COLUMNS].copy()
                    save_df.loc[assign_idx, "실사담당자"] = ""
                    save_df.loc[assign_idx, "실사예정일"] = ""
                    save_df.loc[assign_idx, "진행상태"] = "요청접수"
                    save_inspection_data(save_df)
                    st.cache_data.clear()

                    set_inspection_flash("담당자 배정이 삭제되었습니다.", "success")
                    st.rerun()

    st.divider()

    with st.expander("📝 4. 실사 결과 입력", expanded=False):
        if df.empty:
            st.info("입력할 실사 내역이 없습니다.")
        else:
            result_options = [
                f"{row['row_id']} | {row['현장명']} | {row['실사담당자']} | {row['진행상태']}"
                for _, row in df.iterrows()
            ]

            selected_result = st.selectbox("결과 입력 대상 선택", result_options, key="insp_result_select_new")
            result_idx = int(selected_result.split("|")[0].strip())
            result_row = df.loc[df["row_id"] == result_idx].iloc[0]

            complete_date_raw = str(result_row["실사완료일"]).strip()
            parsed_complete_date = pd.to_datetime(complete_date_raw, errors="coerce")
            default_complete_date = parsed_complete_date.date() if pd.notna(parsed_complete_date) else date.today()

            with st.form(f"inspection_result_form_{result_idx}"):
                r1, r2 = st.columns(2)
                result_text = r1.text_area("실사결과", value=str(result_row["실사결과"]))
                special_note = r2.text_area("특이사항", value=str(result_row["특이사항"]))

                r3, r4 = st.columns(2)
                follow_up = r3.text_area("후속조치", value=str(result_row["후속조치"]))
                complete_date = r4.date_input("실사완료일", value=default_complete_date)

                result_status = st.selectbox(
                    "진행상태",
                    INSPECTION_STATUS_OPTIONS,
                    index=INSPECTION_STATUS_OPTIONS.index(result_row["진행상태"]) if result_row["진행상태"] in INSPECTION_STATUS_OPTIONS else 0
                )

                result_submit = st.form_submit_button("실사 결과 저장")

                if result_submit:
                    save_df = df[INSPECTION_COLUMNS].copy()
                    save_df.loc[result_idx, "실사결과"] = result_text.strip()
                    save_df.loc[result_idx, "특이사항"] = special_note.strip()
                    save_df.loc[result_idx, "후속조치"] = follow_up.strip()
                    save_df.loc[result_idx, "실사완료일"] = str(complete_date)
                    save_df.loc[result_idx, "진행상태"] = result_status
                    save_inspection_data(save_df)

                    set_inspection_flash("실사 결과 저장 완료!", "success")
                    st.rerun()

    st.divider()

    with st.expander("💰 5. 계약 여부 입력", expanded=False):
        if df.empty:
            st.info("계약 처리할 내역이 없습니다.")
        else:
            contract_options = [
                f"{row['row_id']} | {row['현장명']} | {row['상품구분']} | 현재:{row['계약여부']}"
                for _, row in df.iterrows()
            ]

            selected_contract = st.selectbox("계약 처리 대상 선택", contract_options, key="insp_contract_select_new")
            contract_idx = int(selected_contract.split("|")[0].strip())
            contract_row = df.loc[df["row_id"] == contract_idx].iloc[0]

            contract_date_raw = str(contract_row["계약일"]).strip()
            parsed_contract_date = pd.to_datetime(contract_date_raw, errors="coerce")
            default_contract_date = parsed_contract_date.date() if pd.notna(parsed_contract_date) else date.today()

            contract_qty_default = safe_int(contract_row["계약수량"], 0)
            contract_amount_default = safe_int(contract_row["계약금액"], 0)

            with st.form(f"inspection_contract_form_{contract_idx}"):
                ct1, ct2, ct3 = st.columns(3)
                contract_status = ct1.selectbox(
                    "계약여부",
                    CONTRACT_OPTIONS,
                    index=CONTRACT_OPTIONS.index(contract_row["계약여부"]) if contract_row["계약여부"] in CONTRACT_OPTIONS else 0
                )
                contract_date = ct2.date_input("계약일", value=default_contract_date)
                contract_qty = ct3.number_input("계약수량", min_value=0, step=1, value=contract_qty_default)

                ct4, ct5 = st.columns(2)
                contract_amount = ct4.number_input("계약금액", min_value=0, step=10000, value=contract_amount_default)
                fail_reason = ct5.text_input("미계약사유", value=str(contract_row["미계약사유"]))

                contract_submit = st.form_submit_button("계약 정보 저장")

                if contract_submit:
                    save_df = df[INSPECTION_COLUMNS].copy()
                    save_df.loc[contract_idx, "계약여부"] = contract_status
                    save_df.loc[contract_idx, "계약일"] = str(contract_date) if contract_status == "계약" else ""
                    save_df.loc[contract_idx, "계약수량"] = int(contract_qty) if contract_status == "계약" else 0
                    save_df.loc[contract_idx, "계약금액"] = int(contract_amount) if contract_status == "계약" else 0
                    save_df.loc[contract_idx, "미계약사유"] = fail_reason.strip() if contract_status == "미계약" else ""

                    if contract_status == "계약":
                        save_df.loc[contract_idx, "진행상태"] = "계약완료"
                    elif contract_status == "미계약":
                        save_df.loc[contract_idx, "진행상태"] = "미계약종결"

                    save_inspection_data(save_df)

                    set_inspection_flash("계약 정보 저장 완료!", "success")
                    st.rerun()

    st.divider()

    with st.expander("✏️ 6. 상세 보기 / 수정", expanded=False):
        if df.empty:
            st.info("조회할 내역이 없습니다.")
        else:
            view_options = [
                f"{row['row_id']} | {row['현장명']} | {row['상품구분']} | {row['진행상태']}"
                for _, row in df.iterrows()
            ]

            selected_view = st.selectbox("조회 대상 선택", view_options, key="insp_view_select_new")
            view_idx = int(selected_view.split("|")[0].strip())
            view_row = df.loc[df["row_id"] == view_idx].iloc[0]

            is_edit_mode = (
                st.session_state.get("inspection_edit_mode", False)
                and st.session_state.get("inspection_edit_target") == view_idx
            )

            if not is_edit_mode:
                st.markdown("""
                <style>
                .erp-title {
                    font-size: 20px;
                    font-weight: 700;
                    margin-bottom: 14px;
                }
                .erp-summary-wrap {
                    display: grid;
                    grid-template-columns: repeat(4, 1fr);
                    gap: 10px;
                    margin-bottom: 18px;
                }
                .erp-summary-card {
                    border: 1px solid #e5e7eb;
                    border-radius: 10px;
                    padding: 12px 14px;
                    background: #ffffff;
                }
                .erp-summary-label {
                    font-size: 12px;
                    color: #6b7280;
                    margin-bottom: 6px;
                }
                .erp-summary-value {
                    font-size: 20px;
                    font-weight: 600;
                    color: #111827;
                    line-height: 1.2;
                }
                .erp-label {
                    font-size: 13px;
                    font-weight: 600;
                    color: #374151;
                    margin-bottom: 6px;
                }
                .erp-box {
                    border: 1px solid #dbe3ee;
                    border-radius: 10px;
                    background: #eef5ff;
                    padding: 12px 14px;
                    min-height: 46px;
                    font-size: 14px;
                    color: #0f172a;
                    display: flex;
                    align-items: center;
                }
                .erp-textarea {
                    border: 1px solid #e5e7eb;
                    border-radius: 10px;
                    background: #f9fafb;
                    padding: 14px;
                    white-space: pre-wrap;
                    line-height: 1.6;
                    font-size: 14px;
                    color: #111827;
                }
                .erp-section-space {
                    margin-top: 10px;
                    margin-bottom: 10px;
                }
                </style>
                """, unsafe_allow_html=True)

                st.markdown('<div class="erp-title">📄 상세보기</div>', unsafe_allow_html=True)

                st.markdown(
                    f"""
                    <div class="erp-summary-wrap">
                        <div class="erp-summary-card">
                            <div class="erp-summary-label">진행상태</div>
                            <div class="erp-summary-value">{str(view_row["진행상태"])}</div>
                        </div>
                        <div class="erp-summary-card">
                            <div class="erp-summary-label">계약여부</div>
                            <div class="erp-summary-value">{str(view_row["계약여부"])}</div>
                        </div>
                        <div class="erp-summary-card">
                            <div class="erp-summary-label">상품구분</div>
                            <div class="erp-summary-value">{str(view_row["상품구분"])}</div>
                        </div>
                        <div class="erp-summary-card">
                            <div class="erp-summary-label">영업담당자</div>
                            <div class="erp-summary-value">{str(view_row["영업담당자"])}</div>
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

                st.markdown("---")

                a1, a2, a3 = st.columns(3)
                with a1:
                    st.markdown('<div class="erp-label">요청일</div>', unsafe_allow_html=True)
                    st.markdown(f'<div class="erp-box">{str(view_row["요청일"])}</div>', unsafe_allow_html=True)
                with a2:
                    st.markdown('<div class="erp-label">운영사</div>', unsafe_allow_html=True)
                    st.markdown(f'<div class="erp-box">{str(view_row["운영사"])}</div>', unsafe_allow_html=True)
                with a3:
                    st.markdown('<div class="erp-label">현장명</div>', unsafe_allow_html=True)
                    st.markdown(f'<div class="erp-box">{str(view_row["현장명"])}</div>', unsafe_allow_html=True)

                b1, b2, b3 = st.columns(3)
                with b1:
                    st.markdown('<div class="erp-label">현장주소</div>', unsafe_allow_html=True)
                    st.markdown(f'<div class="erp-box">{str(view_row["현장주소"])}</div>', unsafe_allow_html=True)
                with b2:
                    st.markdown('<div class="erp-label">현장연락처</div>', unsafe_allow_html=True)
                    st.markdown(f'<div class="erp-box">{str(view_row["현장연락처"])}</div>', unsafe_allow_html=True)
                with b3:
                    st.markdown('<div class="erp-label">상품구분</div>', unsafe_allow_html=True)
                    st.markdown(f'<div class="erp-box">{str(view_row["상품구분"])}</div>', unsafe_allow_html=True)

                c1, c2, c3 = st.columns(3)
                with c1:
                    st.markdown('<div class="erp-label">주차면수</div>', unsafe_allow_html=True)
                    st.markdown(f'<div class="erp-box">{str(view_row["주차면수"])}</div>', unsafe_allow_html=True)
                with c2:
                    st.markdown('<div class="erp-label">신규설치수량</div>', unsafe_allow_html=True)
                    st.markdown(f'<div class="erp-box">{str(view_row["신규설치수량"])}</div>', unsafe_allow_html=True)
                with c3:
                    st.markdown('<div class="erp-label">기설치수량</div>', unsafe_allow_html=True)
                    st.markdown(f'<div class="erp-box">{str(view_row["기설치수량"])}</div>', unsafe_allow_html=True)

                d1, d2 = st.columns(2)
                with d1:
                    st.markdown('<div class="erp-label">영업담당자</div>', unsafe_allow_html=True)
                    st.markdown(f'<div class="erp-box">{str(view_row["영업담당자"])}</div>', unsafe_allow_html=True)
                with d2:
                    st.markdown('<div class="erp-label">영업담당자 연락처</div>', unsafe_allow_html=True)
                    st.markdown(f'<div class="erp-box">{str(view_row["영업담당연락처"])}</div>', unsafe_allow_html=True)

                st.markdown('<div class="erp-section-space"></div>', unsafe_allow_html=True)

                st.markdown('<div class="erp-label">요청내용</div>', unsafe_allow_html=True)
                request_text = str(view_row["요청내용"]).strip()
                if request_text:
                    st.markdown(
                        f'<div class="erp-textarea">{request_text}</div>',
                        unsafe_allow_html=True
                    )
                else:
                    st.markdown('<div class="erp-box">요청내용이 없습니다.</div>', unsafe_allow_html=True)

                st.markdown('<div class="erp-section-space"></div>', unsafe_allow_html=True)

                st.markdown('<div class="erp-label">비고</div>', unsafe_allow_html=True)
                note_text = str(view_row["비고"]).strip()
                if note_text:
                    st.markdown(
                        f'<div class="erp-textarea">{note_text}</div>',
                        unsafe_allow_html=True
                    )
                else:
                    st.markdown('<div class="erp-box">비고가 없습니다.</div>', unsafe_allow_html=True)

                st.markdown('<div class="erp-section-space"></div>', unsafe_allow_html=True)

                st.markdown('<div class="erp-label">첨부파일</div>', unsafe_allow_html=True)
                if str(view_row["첨부파일링크"]).strip():
                    file_name = str(view_row["첨부파일명"]).strip() or "첨부파일 열기"
                    st.link_button(file_name, str(view_row["첨부파일링크"]))
                else:
                    st.markdown('<div class="erp-box">첨부파일이 없습니다.</div>', unsafe_allow_html=True)

                st.markdown("---")

                if st.button("수정", use_container_width=True, key=f"insp_edit_mode_btn_{view_idx}"):
                    st.session_state.inspection_edit_mode = True
                    st.session_state.inspection_edit_target = view_idx
                    st.rerun()

            else:
                st.markdown("## ✏️ 수정 모드")

                req_date_raw = str(view_row["요청일"]).strip()
                parsed_req_date = pd.to_datetime(req_date_raw, errors="coerce")
                default_req_date = parsed_req_date.date() if pd.notna(parsed_req_date) else date.today()

                with st.form(f"inspection_edit_form_{view_idx}"):
                    e1, e2, e3 = st.columns(3)
                    edit_req_date = e1.date_input("요청일 수정", value=default_req_date)
                    edit_operator = e2.text_input("운영사 수정", value=str(view_row["운영사"]))
                    edit_name = e3.text_input("현장명 수정", value=str(view_row["현장명"]))

                    e4, e5, e6 = st.columns(3)
                    edit_addr = e4.text_input("현장주소 수정", value=str(view_row["현장주소"]))
                    edit_phone = e5.text_input("현장연락처 수정", value=str(view_row["현장연락처"]))
                    edit_product = e6.selectbox(
                        "상품구분 수정",
                        PRODUCT_OPTIONS,
                        index=PRODUCT_OPTIONS.index(view_row["상품구분"]) if view_row["상품구분"] in PRODUCT_OPTIONS else 0
                    )

                    e7, e8, e9 = st.columns(3)
                    edit_parking = e7.number_input("주차면수 수정", min_value=0, step=1, value=safe_int(view_row["주차면수"], 0))
                    edit_new_qty = e8.number_input("신규설치수량 수정", min_value=0, step=1, value=safe_int(view_row["신규설치수량"], 0))
                    edit_old_qty = e9.number_input("기설치수량 수정", min_value=0, step=1, value=safe_int(view_row["기설치수량"], 0))

                    e10, e11 = st.columns(2)
                    edit_sales = e10.text_input("영업담당자 수정", value=str(view_row["영업담당자"]))
                    edit_sales_phone = e11.text_input("영업담당자 연락처 수정", value=str(view_row["영업담당연락처"]))

                    edit_request = st.text_area("요청내용 수정", value=str(view_row["요청내용"]))
                    edit_note = st.text_input("비고 수정", value=str(view_row["비고"]))

                    s1, s2 = st.columns(2)
                    save_submit = s1.form_submit_button("기본 정보 수정 저장", use_container_width=True)
                    cancel_submit = s2.form_submit_button("취소", use_container_width=True)

                    if save_submit:
                        save_df = df[INSPECTION_COLUMNS].copy()
                        save_df.loc[view_idx, "요청일"] = str(edit_req_date)
                        save_df.loc[view_idx, "운영사"] = edit_operator.strip()
                        save_df.loc[view_idx, "현장명"] = edit_name.strip()
                        save_df.loc[view_idx, "현장주소"] = edit_addr.strip()
                        save_df.loc[view_idx, "현장연락처"] = edit_phone.strip()
                        save_df.loc[view_idx, "상품구분"] = edit_product
                        save_df.loc[view_idx, "주차면수"] = int(edit_parking)
                        save_df.loc[view_idx, "신규설치수량"] = int(edit_new_qty)
                        save_df.loc[view_idx, "기설치수량"] = int(edit_old_qty)
                        save_df.loc[view_idx, "영업담당자"] = edit_sales.strip()
                        save_df.loc[view_idx, "영업담당연락처"] = edit_sales_phone.strip()
                        save_df.loc[view_idx, "요청내용"] = edit_request.strip()
                        save_df.loc[view_idx, "비고"] = edit_note.strip()
                        save_inspection_data(save_df)

                        st.session_state.inspection_edit_mode = False
                        st.session_state.inspection_edit_target = None
                        set_inspection_flash("기본 정보 수정 완료!", "success")
                        st.rerun()

                    if cancel_submit:
                        st.session_state.inspection_edit_mode = False
                        st.session_state.inspection_edit_target = None
                        st.rerun()

    st.divider()

    with st.expander("🗑️ 7. 실사 요청 삭제", expanded=False):
        if df.empty:
            st.info("삭제할 내역이 없습니다.")
        else:
            delete_options = [
                f"{row['row_id']} | {row['현장명']} | {row['상품구분']} | {row['영업담당자']}"
                for _, row in df.iterrows()
            ]

            selected_delete = st.selectbox("삭제할 내역 선택", delete_options, key="insp_delete_select_new")
            confirm_delete = st.checkbox("정말 삭제합니다. 되돌리기 어렵습니다.", key="insp_delete_confirm_new")

            if st.button("선택 내역 삭제", key="insp_delete_btn_new"):
                if not confirm_delete:
                    st.warning("삭제 확인 체크를 먼저 해주세요.")
                else:
                    delete_idx = int(selected_delete.split("|")[0].strip())
                    save_df = df[INSPECTION_COLUMNS].copy()
                    save_df = save_df.drop(index=delete_idx).reset_index(drop=True)
                    save_inspection_data(save_df)

                    set_inspection_flash("삭제가 완료되었습니다.", "success")
                    st.rerun()


# =========================================================
# 6. 메인 실행
# =========================================================
if not st.session_state.logged_in:
    login_screen()
else:
    draw_sidebar()

    if st.session_state.menu == "홈":
        home_page()
    elif st.session_state.menu == "연차 관리":
        vacation_page()
    elif st.session_state.menu == "시공 일정":
        schedule_page()
    elif st.session_state.menu == "실사 관리":
        inspection_page()
